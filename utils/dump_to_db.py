# coding: utf-8
import pymysql
import openpyxl


class Dumper(object):
    def __init__(self):
        self.conn = pymysql.Connection(
            host="gz-cdb-h25tz5ek.sql.tencentcdb.com",
            user="weibo",
            password="weibo_QWer",
            database="poem",
            port=62581,
            charset="utf8"
        )
        self.wb = None

    def get_poet_id(self, name):
        cur = self.conn.cursor()
        sql = "SELECT id FROM poet WHERE name='{}'".format(name)
        cur.execute(sql)
        res = cur.fetchone()
        if not res:
            return None
        return res[0]

    def dump_poem_to_file(self):
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM poem")
        count = cur.fetchone()[0]
        print(count)
        LIMIT = 10000
        OFFSET = 0
        while OFFSET < count:
            sql = "SELECT content from poem LIMIT {} OFFSET {} ".format(LIMIT, OFFSET)
            cur.execute(sql)
            for content in cur.fetchall():
                print(content)

    def read_excel(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        sheet = self.wb.active
        for row in sheet.iter_rows(min_row=2):
            yield {
                "title": row[2].value,
                "content": row[3].value,
                "yunlv_rule": "NULL",
                "poem_type": "poem",
                "author": row[4].value,
            }

    def read_poetry_excel(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        sheet = self.wb.active
        for row in sheet.iter_rows(min_row=2):
            yield {
                "title": row[2].value,
                "content": row[3].value,
                "yunlv_rule": row[4].value,
                "poem_type": "poetry",
                "author": row[5].value,
            }

    def read_poetry(self, filename, min_row, max_row):
        self.wb = openpyxl.load_workbook(filename)
        sheet = self.wb.active
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row):
            yield {
                "title": row[2].value,
                "content": row[3].value,
                "yunlv_rule": row[4].value,
                "poem_type": "poetry",
                "author": row[5].value,
            }

    def dump(self, filename, read_func, output_name):
        sql = """
        INSERT INTO poem(title, content, yunlv_rule, poem_type, author_id) values
        """
        author_cache = {}
        for poem in read_func(filename):
            if not author_cache.get(poem.get('author')):
                author_id = self.get_poet_id(poem.get('author'))
                author_cache[poem.get('author')] = author_id
            else:
                author_id = author_cache.get(poem.get('author'))
            poem.pop('author')
            if not author_id:
                continue
            poem['author_id'] = author_id
            s = "('{title}', '{content}', '{yunlv_rule}', '{poem_type}', {author_id}),".format(
                **poem
            )
            sql += s
        with open(output_name, 'w', encoding='utf-8') as f:
            f.write(sql)



if __name__ == '__main__':
    d = Dumper()
    d.dump_poem_to_file()
    # poem_path = "C:/Users/lzj/Desktop/poetry.xlsx"
    # # print(d.get_poet_id('苏'))
    # d.dump(poem_path, d.read_poetry_excel, 'poetry.sql')


